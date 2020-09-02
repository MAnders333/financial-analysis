import os
import sys
import openpyxl
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


class FinancialAnalyzer:
    """
    This class represents an object that checks whether a given Excel file for a monthly financial
    analysis is up to date. If no Excel file exists, it creates one in the current directory. 

    :param path: path to an existing Excel file for the monthly analysis. If no Excel file exists, do not specify path and an Excel file will be created.
    :param sheet_name: name of the sheet in the Excel file used for the monthly financial analysis. 
    :type path: string
    :type sheet_name: string
    """

    def __init__(self, path=None, sheet_name="Financial Overview"):
        """
        Initializes the FinancialAnalyzer object. During initializing, it checks whether the Excel file is up to date.
        """
        # Sets name for the sheet
        self.sheet_name = sheet_name

        # Checks whether a path is specified
        if path != None:
            # If a path is specified, path attribute is set
            self.path = str(path)

            # Tries to open Excel file and worksheet
            try:
                self.book = openpyxl.load_workbook(self.path)
            except FileNotFoundError:
                print("File could not be found. Please provide a valid path.")
            try:
                self.sheet = self.book[sheet_name]
            except:
                print("Worksheet could not be found. Please provide a valid sheet name.")

            # Checks whether worksheet is up to date
            self.last_row = self.sheet.max_row
            self.current_month = datetime.date(datetime.now()).month
            if int(self.sheet.cell(self.last_row, 2).value) != int(self.current_month):
                self.isUpdated = False
                print("Worksheet needs to be updated.")
            else:
                self.isUpdated = True
                print("Worksheet is up to date.")
        else:
            # If path is not specified, a new Excel file and worksheet are created
            self.book = openpyxl.Workbook()
            print("Workbook created.")
            self.sheet = self.book.create_sheet(title=sheet_name, index=0)
            print("Worksheet created.")
            headers = ["year", "month", "cash", "cashflow", "bank", "bank_cashflow", "investments",
                       "investments_total_change", "assets_total", "assets_total_change", "assets_rel_change"]
            for i, header in enumerate(headers):
                self.sheet.cell(1, i+1).value = header
            self.book.save("monthly_budget.xlsx")

            # As a new file and sheet were created, it needs to be updated with the account balances of the current month
            self.isUpdated = False
            print("Worksheet needs to be updated.")

    def scrape_bank_website(self, url, username, password):
        """
        This method logs into your online banking account and scrapes the account balances from the website.

        !!!It needs to be adapted for your bank website as its layout will be different!!!

        :param url: url of your bank's login page
        :param username: username of your online banking account
        :param password: password of your online banking account
        :type url: string
        :type username: string
        :type password: string

        For security reasons, it is recommended to use environment variables for your username and password.
        """

        # Visits the bank website
        self.driver = webdriver.Firefox()
        self.driver.get(url)

        # Waits for a specified element to render to make sure the page is fully loaded
        WebDriverWait(self.driver, 15).until(EC.presence_of_element_located(
            (By.XPATH, """//*[@id="txtBenutzerkennung"]""")))  # Change the xpath according to the element's xpath of your banks website

        # Writes the username and password into the login form
        element_username = self.driver.find_element_by_xpath(
            """//*[@id="txtBenutzerkennung"]""")
        element_username.send_keys(username)
        element_password = self.driver.find_element_by_xpath(
            """//*[@id="pwdPin"]""")
        element_password.send_keys(password)

        # Clicks the login button
        login_button = self.driver.find_element_by_xpath(
            """//*[@id="xview-anmelden"]""")
        login_button.click()

        # Waits for a specified element to render to make sure the page is fully loaded
        WebDriverWait(self.driver, 15).until(EC.presence_of_element_located(
            (By.XPATH, """/html/body/div[2]/div/div[1]/div[3]/div[1]/div[3]/div[2]/div/table/tbody/tr/td/div/table/tbody/tr[1]/td/span/div/div/div[2]/div/div/div/div[3]/div[2]/div[2]/div/div[3]/div/div/div/div[3]/div/div/table/tbody/tr/td[3]/div/div""")))

        # Reads out balances in the bank accounts
        self.balance = float(self.driver.find_element_by_xpath(
            """/html/body/div[2]/div/div[1]/div[3]/div[1]/div[3]/div[2]/div/table/tbody/tr/td/div/table/tbody/tr[1]/td/span/div/div/div[2]/div/div/div/div[3]/div[2]/div[2]/div/div[3]/div/div/div/div[3]/div/div/table/tbody/tr/td[3]/div/div""").text.strip().split(" ")[0].replace(".", "").replace(",", "."))

        # Quits browser
        self.driver.quit()

    def update_workbook(self, cash=0.0):
        """
        This method updates the Excel file with the current bank balances scraped from the website plus the amount of cash in your wallet.

        :param cash: the amount of cash in your wallet
        :type cash: integer / float
        """

        # Sets cash attribute
        self.cash = float(cash)

        current_year = datetime.date(datetime.now()).year
        bank_balance = self.balance
        investments = 0  # Fixed rate of investments each month
        assets_total = self.cash + bank_balance + investments

        # Checks whether this is the first entry (so cash flow and other changes in assets cannot be calculated)
        if self.last_row == 1:
            cashflow, bank_cashflow, investments_change, assets_total_change, assets_rel_change = 0
        else:
            cashflow = self.cash - \
                float(self.sheet.cell(self.last_row, 3).value)
            bank_cashflow = bank_balance - \
                float(self.sheet.cell(self.last_row, 5).value)
            investments_change = investments - \
                float(self.sheet.cell(self.last_row, 7).value)
            assets_total_change = assets_total - \
                float(self.sheet.cell(self.last_row, 9).value)
            assets_rel_change = assets_total_change / \
                float(self.sheet.cell(self.last_row, 9).value)

        # Adds new row of values into the Excel file
        current_values = [current_year, self.current_month, self.cash, cashflow, bank_balance, bank_cashflow,
                          investments, investments_change, assets_total, assets_total_change, assets_rel_change]
        for i, value in enumerate(current_values):
            self.sheet.cell(self.last_row + 1, i + 1).value = value

        # Saves Excel file
        self.book.save(self.path)

        print("Worksheet successfully updated.")
        print(f"""
        cash: {round(self.cash, 2)}
        bank: {round(bank_balance, 2)}
        investments: {round(investments, 2)}
        cashflow: {round(cashflow, 2)}
        bank cashflow: {round(bank_cashflow, 2)}

        total assets: {round(assets_total, 2)}
        total assets change: {round(assets_total_change, 2)}
        total assets relative change: {round(assets_rel_change, 2) * 100}
        """)


if __name__ == "__main__":
    # Add the path to your Excel file here
    path = "/Users/marcanders/Projects/financial-analysis/monthly_budget.xlsx"
    url = os.environ.get("BK_URL")  # Add your bank's url here
    username = os.environ.get("BK_UN")  # Add your bank's username here
    password = os.environ.get("BK_PW")  # Add your bank's password here
    if len(sys.argv) > 1:
        # Specify your current cash balance in the terminal as an argument
        cash = sys.argv[1]
    else:
        cash = 0.0

    analyzer = FinancialAnalyzer(path=path)
    if analyzer.isUpdated == False:
        analyzer.scrape_bank_website(url, username, password)
        analyzer.update_workbook(cash)
